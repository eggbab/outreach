#!/usr/bin/env python3
"""
=============================================================
마트마트 광고주 자동 수집 & 이메일/DM 발송 올인원 프로그램
=============================================================

★ 브라우저 안 띄움! 100% requests 기반 (백그라운드 실행)
★ 네이버 검색 + 네이버 쇼핑 + 네이버 지도 + 구글 + 인스타 통합
★ 1000개 이상 수집 목표

실행:
  pip3 install openpyxl beautifulsoup4 requests instagrapi

  python3 마트마트_자동화.py               # 전체 실행 (수집→요약→승인→발송)
  python3 마트마트_자동화.py collect        # 업체 수집만
  python3 마트마트_자동화.py collect-insta  # 인스타만 수집
  python3 마트마트_자동화.py send-test      # 테스트 이메일
  python3 마트마트_자동화.py status         # 현황 확인
=============================================================
"""

import os
import re
import sys
import json
import time
import random
import smtplib
import warnings
from datetime import datetime
from urllib.parse import urlparse, urljoin, quote
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders

# SSL 경고 억제
warnings.filterwarnings('ignore', message='Unverified HTTPS request')
try:
    import urllib3
    urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
except ImportError:
    pass

# ============================================================
#  설정
# ============================================================
GMAIL_EMAIL = "gimuuuujin@gmail.com"
GMAIL_APP_PASSWORD = "xaqu biuk suaj tort"
_BASE_DIR = os.path.dirname(os.path.abspath(__file__))
EXCEL_FILE = os.path.join(_BASE_DIR, "마트마트_잠재광고주_리스트.xlsx")
PROGRESS_FILE = os.path.join(_BASE_DIR, "수집_진행상황.json")
EMAIL_DAILY_FILE = os.path.join(_BASE_DIR, "email_일일발송.json")

# 이메일 설정
DAILY_SEND_LIMIT = 80
DELAY_BETWEEN_EMAILS = (25, 45)

# 인스타그램 설정
INSTA_USERNAME = "e99bab"
INSTA_PASSWORD = "apple6627kwJ^"
INSTA_SESSION_FILE = "insta_session.json"
INSTA_DM_PROGRESS_FILE = "dm_진행상황.json"
DAILY_DM_LIMIT = 15
DELAY_BETWEEN_DMS = (90, 180)

# ============================================================
#  검색 키워드 (대폭 확장 — 1000개+ 수집 목표)
# ============================================================
KEYWORDS = [
    # 냉동냉장 설비 (뉴스 안 나오는 구체적 제품 키워드)
    "업소용 냉장고", "냉동쇼케이스", "오픈쇼케이스", "업소용 냉동고",
    "냉장쇼케이스 제조", "업소용 제빙기", "산업용 냉장고",
    "업소용 김치냉장고", "쇼케이스 냉장고 도매", "저온저장고",
    # 매장 인테리어/설비
    "마트 진열대 곤돌라", "마트 간판 제작", "LED 전광판 마트",
    "매장 인테리어 시공", "슈퍼마켓 리모델링", "마트 조명 LED",
    "곤돌라 진열장 제조", "상업용 조명 설치",
    # POS/결제
    "업소용 POS", "키오스크 셀프계산대", "카드단말기 업소용",
    "마트 POS시스템", "무인결제 키오스크", "바코드 스캐너 업소용",
    # 보안
    "마트 CCTV 보안", "도난방지 보안태그", "EAS 도난방지시스템",
    "업소용 CCTV 설치", "매장 도난방지 게이트",
    # 계량/라벨
    "전자저울 업소용", "바코드프린터 라벨프린터", "가격표시기 전자라벨",
    "라벨 인쇄기 업소용", "전자가격표시기",
    # 포장
    "업소용 진공포장기", "식품포장용기 도매", "비닐봉투 쇼핑백 제조",
    "랩포장기 업소용", "식품 트레이 도매", "포장재 도매업체",
    "스티로폼 박스 도매", "진공포장비닐 도매",
    # 위생/청소
    "업소용 청소장비", "마트 앞치마 위생복", "업소용 위생용품",
    "산업용 청소기", "업소용 세제 도매",
    # 식자재/도매 (업체명 포함 키워드 → 뉴스 배제)
    "식자재 도매 납품", "수입식품 도매 유통",
    "과일 도매 납품", "정육 도매 납품업체", "수산물 도매 업체",
    "냉동식품 도매", "음료 도매 납품", "과자 도매 유통",
    "조미료 도매", "가공식품 도매", "유제품 도매 납품",
    "농산물 도매 납품", "반찬 도매 업체",
    # 기타 설비
    "업소용 정수기 렌탈", "해충방제 방역 업소", "마트 쇼핑카트",
    "음식물처리기 업소용", "업소용 에어컨 냉난방", "상업용 정수기",
    # 포장재/소모품
    "장갑 위생장갑 도매", "일회용품 도매", "위생모 도매",
    "영수증 용지 도매", "쇼핑백 제작",
    # 인쇄/홍보물 (마트 전단지/현수막)
    "전단지 인쇄 업체", "현수막 제작",

    # ── 냉동냉장 추가 변형 ──
    "냉동쇼케이스 도매", "업소용 냉동고 판매", "정육쇼케이스 업체",
    "오픈쇼케이스 냉장고", "워크인쿨러 설치", "냉장 진열장 제조",
    "냉동창고 설비업체", "쇼케이스 냉장고 제조사", "냉동탑차 설비",
    "업소용 냉장고 대리점", "냉동냉장 설비 전문", "저온 창고 시공",
    "수산물 냉장쇼케이스", "아이스크림 냉동고 업소용", "냉장고 도매상",

    # ── 주방 설비 추가 ──
    "황학동 주방기기", "업소용 식기세척기", "업소용 가스레인지",
    "업소용 가스레인지 판매", "주방설비 전문업체", "업소용 인덕션 레인지",
    "업소용 오븐 판매", "업소용 튀김기", "업소용 솥 도매",
    "스테인리스 주방기구", "업소용 작업대 제조", "주방기기 도매상",
    "업소용 조리기구 납품", "식당 주방설비 시공", "업소용 주방용품 총판",
    "황학동 주방설비 업체", "업소용 주방 기구 대리점",

    # ── 마트 집기 / 진열대 추가 ──
    "마트 집기 도매", "슈퍼마켓 집기", "곤돌라 진열대 제조",
    "마트 진열대 판매", "마트 바구니 쇼핑카트 도매", "진열집기 전문",
    "매장 진열장 납품", "소매점 집기 도매", "마트 선반 제조",
    "편의점 집기 납품", "슈퍼마켓 인테리어 집기", "진열대 도매상",

    # ── POS / IT 추가 ──
    "마트 포스기 대리점", "편의점 POS 납품", "POS단말기 전문업체",
    "무인매장 솔루션", "재고관리 프로그램 업체", "매장관리 소프트웨어",
    "상품 발주 시스템 업체", "바코드 프린터 납품", "라벨 프린터 업소용",
    "전자가격표시기 업체", "무인 셀프계산대", "키오스크 제조사",
    "마트 키오스크 납품", "편의점 무인결제", "POS 솔루션 업체",

    # ── 포장 / 기계 추가 ──
    "식품포장기 업체", "비닐봉투 도매상", "포장비닐 제조",
    "식품트레이 도매", "스티로폼 용기 도매", "포장용기 도매상",
    "랩포장기 업소", "진공포장기 업소용", "자동포장기 제조",
    "쇼핑백 도매상", "육류포장 트레이", "PET용기 도매",
    "일회용 용기 도매", "뚜껑 용기 포장재", "포장재 납품업체",
    "스트레치필름 도매", "에어캡 뽁뽁이 도매", "택배박스 도매",
    "식품 밀봉기 업체", "열수축필름 도매",

    # ── 식자재 도매 확장 ──
    "서울 식자재 도매", "인천 식자재 납품", "부산 식자재 도매",
    "경기 식자재 납품", "대구 식자재 도매", "대전 식자재 납품",
    "광주 식자재 도매", "수도권 식자재 납품업체",
    "수산물 도매상", "냉동수산물 도매", "건어물 도매업체",
    "쌀 도매 납품", "잡곡 도매", "계란 도매납품",
    "유제품 도매 납품", "치즈 버터 도매", "두부 콩나물 납품",
    "반찬 도매 납품", "가공식품 납품업체", "HMR 간편식 도매",
    "수입과일 도매", "국내산 과일 도매납품", "채소 도매상",
    "냉동야채 도매", "버섯 도매 납품", "김치 도매 납품",
    "젓갈 반찬 도매", "장류 도매납품", "조미료 도매상",

    # ── 음료 / 주류 도매 ──
    "음료 도매 납품업체", "탄산음료 도매", "생수 도매 납품",
    "주류 도매상", "맥주 도매 납품", "소주 도매",
    "막걸리 도매납품", "와인 도매상", "커피 도매 납품",
    "에너지음료 도매", "식혜 음료 도매", "과일주스 도매납품",

    # ── 인테리어 / 간판 확장 ──
    "마트 인테리어 전문", "슈퍼마켓 인테리어 업체", "소매점 인테리어",
    "상업용 LED 조명", "매장 조명 업체", "LED 간판 제작",
    "아크릴 간판 제작", "현수막 인쇄업체", "배너 현수막 도매",
    "전단지 인쇄 도매", "스티커 라벨 인쇄", "쇼핑백 인쇄 제작",
    "어닝 차양 설치", "롤스크린 업소용", "매장 사인물 제작",

    # ── 청소 / 위생 확장 ──
    "업소용 세제 납품", "주방세제 도매", "청소용품 도매상",
    "위생복 앞치마 도매", "일회용 장갑 도매", "위생모 마스크 도매",
    "걸레 청소포 도매", "쓰레기봉투 도매", "방역소독 업소",
    "해충방제 전문업체", "업소용 핸드드라이어", "자동손소독기 업소",

    # ── 보안 / 계량 확장 ──
    "마트 CCTV 설치업체", "매장 보안 시스템", "도난방지 태그 납품",
    "EAS 보안태그 도매", "도난방지 게이트 설치", "산업용 저울 납품",
    "전자저울 도매상", "업소용 계량기 대리점",

    # ── 기타 설비 확장 ──
    "업소용 정수기 납품", "업소용 에어컨 업체", "상업용 공기청정기",
    "음식물처리기 업소", "업소용 음식물분쇄기", "자동판매기 납품",
    "무인자판기 설치", "영수증 용지 도매", "POS 소모품 도매",
    "매장 음향기기 납품", "업소용 냉온수기",

    # ── 지역별 특화 ──
    "서울 마트 설비업체", "경기 마트 납품업체", "인천 마트 설비",
    "부산 냉동설비 업체", "대구 주방기기 업체", "광주 식자재 납품",
    "대전 포장재 도매", "충남 마트 납품", "경남 식자재 도매",
    "강원 식자재 납품", "전북 농산물 도매", "제주 식자재 납품",
]

# 네이버 쇼핑 전용 키워드 (smartstore.naver.com만 수집 → 뉴스기사 0%)
# 110개 → 키워드당 40개 = 최대 4,400개 스토어 수집 가능
SHOPPING_KEYWORDS = [
    # ── 냉동냉장 설비 (20) ──
    "업소용냉장고", "냉동쇼케이스", "오픈쇼케이스", "업소용냉동고",
    "업소용제빙기", "냉장쇼케이스", "워크인냉장고", "저온저장고",
    "아이스크림냉동고", "수산물냉장고", "정육쇼케이스", "냉동진열장",
    "업소용김치냉장고", "냉동창고설비", "냉장고쇼케이스도매",
    "냉동탑차냉장설비", "냉장이동카트", "업소용소형냉동고",
    "냉장진열대", "오픈냉장쇼케이스",
    # ── 주방·조리 설비 (12) ──
    "업소용주방기기", "업소용가스레인지", "업소용식기세척기",
    "황학동주방설비", "업소용인덕션", "업소용솥",
    "업소용조리대", "스테인리스주방", "업소용튀김기",
    "업소용오븐", "업소용작업대", "업소용칼",
    # ── 마트 집기·인테리어 (15) ──
    "곤돌라진열대", "마트진열대", "마트선반", "마트집기",
    "진열집기", "마트쇼케이스", "슈퍼마켓진열대", "마트바구니",
    "쇼핑카트", "마트쇼핑카트", "매장인테리어", "상업용조명",
    "업소용조명", "LED간판", "마트간판",
    # ── POS·결제·IT (15) ──
    "업소용POS", "마트POS", "편의점POS", "바코드스캐너",
    "가격표시기", "전자가격표", "무인결제키오스크", "셀프계산대",
    "카드단말기", "바코드프린터", "라벨프린터", "영수증프린터",
    "키오스크", "무인매장솔루션", "재고관리프로그램",
    # ── 보안·계량 (10) ──
    "마트CCTV", "업소용CCTV", "도난방지게이트", "EAS도난방지",
    "도난방지태그", "도난방지스티커", "전자저울업소용",
    "산업용저울", "업소용계량기", "정밀전자저울",
    # ── 포장 기계·포장재 (15) ──
    "업소용진공포장기", "랩핑기업소", "식품밀봉기",
    "진공포장비닐", "스트레치필름", "자동포장기",
    "식품포장용기", "식품트레이", "비닐봉투도매",
    "쇼핑백도매", "포장비닐도매", "일회용포장용기",
    "스티로폼박스도매", "육류트레이", "포장재도매",
    # ── 청소·위생·소모품 (12) ──
    "업소용청소기", "업소용세제", "위생복도매", "앞치마도매",
    "업소용고무장갑", "위생모", "방역용품", "일회용품도매",
    "장갑도매", "마스크도매업소", "쓰레기봉투도매", "걸레도매",
    # ── 식자재·도매 (20) ──
    "식자재도매", "정육도매", "수산물도매", "과일도매",
    "채소도매", "냉동식품도매", "가공식품도매", "건어물도매",
    "젓갈도매", "음료도매", "주류도매", "조미료도매",
    "유제품도매", "계란도매", "쌀도매", "잡곡도매",
    "통조림도매", "라면도매", "과자도매", "수입식품도매",
    # ── 기타 설비·서비스 (10) ──
    "업소용정수기", "음식물처리기", "해충방제", "업소용에어컨",
    "음식물분쇄기", "업소용핸드드라이어", "영수증용지",
    "쇼핑백제작", "현수막제작", "전단지인쇄",
]

# 네이버 지도 키워드
MAP_KEYWORDS = [
    "마트 설비 업체", "냉동 설비 업체", "업소용 주방기기",
    "매장 인테리어 업체", "POS 설치 업체", "CCTV 설치 업체",
    "간판 제작 업체", "식자재 도매 매장", "포장재 도매 매장",
    "청소용품 도매", "냉동창고 업체", "물류 업체",
]

# 인스타그램 해시태그
INSTA_HASHTAGS = [
    "업소용냉장고", "냉동쇼케이스", "마트인테리어", "매장인테리어",
    "업소용POS", "키오스크", "셀프계산대", "마트설비",
    "진열대", "곤돌라진열대", "마트창업", "슈퍼마켓창업",
    "식자재도매", "업소용제빙기", "냉동냉장", "마트리모델링",
    "업소용가구", "카드단말기", "전자저울", "마트간판",
    "식품포장", "업소용냉동고", "콜드체인", "마트CCTV",
    "매장리모델링", "업소용주방", "도매시장", "식자재납품",
    "냉장쇼케이스", "무인매장", "무인결제", "마트컨설팅",
    "유통업", "소매업", "프랜차이즈마트", "편의점창업",
]

# 인스타그램 키워드 검색용
INSTA_SEARCH_KEYWORDS = [
    "업소용 냉장고", "마트 설비", "매장 인테리어", "식자재 도매",
    "냉동 쇼케이스", "마트 POS", "키오스크", "진열대",
    "냉동냉장 설비", "업소용 주방", "무인매장", "유통 솔루션",
    "마트 창업", "편의점 설비", "식품 포장", "간판 제작",
]


# ============================================================
#  인스타그램 DM 메시지 템플릿
# ============================================================
DM_TEMPLATES = [
    (
        "안녕하세요 {company}님!\n\n"
        "저는 네이버 카페 '마트마트' 대표 김우진입니다.\n"
        "마트·슈퍼마켓 업계 종사자 22,000명이 모인 커뮤니티인데, "
        "귀사 제품·서비스를 저희 회원분들께 소개해드리면 좋을 것 같아 연락드렸습니다.\n\n"
        "광고 상품은 PC·모바일 배너, 전용 홍보 게시판, 공동구매 프로모션, 전체 회원 이메일 발송 등이 있고 "
        "업종별 독점으로 운영됩니다.\n\n"
        "관심 있으시면 편하게 답장 주세요. 자세한 제안서 바로 보내드리겠습니다!\n"
        "카페: https://cafe.naver.com/martmart"
    ),
]


# ============================================================
#  1단계: 업체 수집 (100% requests — 브라우저 없음)
# ============================================================
class BusinessCollector:
    """네이버 + 구글 + 네이버 쇼핑 + 네이버 지도 업체 수집 (브라우저 안 씀)"""

    def __init__(self):
        self.collected = {}
        self._http = None
        self.done_keywords = set()
        self.done_shopping = set()
        self.done_map = set()
        self.load_progress()

    # ── 진행상황 ──
    def load_progress(self):
        if os.path.exists(PROGRESS_FILE):
            try:
                with open(PROGRESS_FILE, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                    self.collected = data.get('collected', {})
                    self.done_keywords = set(data.get('done_keywords', []))
                    self.done_shopping = set(data.get('done_shopping', []))
                    self.done_map = set(data.get('done_map', []))
                    print(f"  이전 진행상황 로드: {len(self.collected)}개 업체, {len(self.done_keywords)}개 키워드 완료")
            except Exception:
                self.collected = {}
                self.done_keywords = set()
                self.done_shopping = set()
                self.done_map = set()
        else:
            pass

    def save_progress(self):
        with open(PROGRESS_FILE, 'w', encoding='utf-8') as f:
            json.dump({
                'collected': self.collected,
                'done_keywords': list(self.done_keywords),
                'done_shopping': list(self.done_shopping),
                'done_map': list(self.done_map),
                'last_update': datetime.now().isoformat(),
            }, f, ensure_ascii=False, indent=2)

    # ── HTTP 세션 ──
    def _get_http(self):
        if self._http is None:
            import requests
            self._http = requests.Session()
            self._http.headers.update({
                'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) '
                              'AppleWebKit/537.36 (KHTML, like Gecko) '
                              'Chrome/124.0.0.0 Safari/537.36',
                'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8',
                'Accept-Language': 'ko-KR,ko;q=0.9,en;q=0.8',
            })
            self._http.verify = False
        return self._http

    def _http_get(self, url, timeout=10, headers=None):
        try:
            kw = {'timeout': timeout, 'allow_redirects': True}
            if headers:
                kw['headers'] = headers
            resp = self._get_http().get(url, **kw)
            resp.encoding = resp.apparent_encoding or 'utf-8'
            if resp.status_code == 200:
                return resp.text
        except Exception:
            pass
        return ''

    def _add_business(self, domain_key, info):
        """중복 체크 후 업체 추가"""
        if domain_key in self.collected:
            return False
        skip_domains = ['naver.com', 'google.com', 'youtube.com', 'facebook.com',
                        'instagram.com', 'twitter.com', 'tistory.com', 'blog.naver',
                        'cafe.naver', 'wikipedia', 'namu.wiki', 'daum.net',
                        'dcinside', 'clien', 'reddit', 'kin.naver', 'shopping.naver',
                        'search.naver', 'map.naver']
        if any(s in domain_key for s in skip_domains):
            # smartstore는 허용
            if 'smartstore' not in domain_key:
                return False
        self.collected[domain_key] = info
        return True

    # ── [1] 네이버 검색 (requests) ──
    def search_naver(self, keyword):
        """네이버 통합검색 — requests로 파워링크 + 일반 결과 수집"""
        from bs4 import BeautifulSoup
        count = 0

        # 네이버 검색 API가 아닌 웹 검색 페이지 직접 크롤링
        url = f"https://search.naver.com/search.naver?query={quote(keyword)}&where=web"
        headers = {
            'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) '
                          'AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36',
            'Referer': 'https://www.naver.com/',
        }
        html = self._http_get(url, headers=headers)
        if not html:
            return count

        soup = BeautifulSoup(html, 'html.parser')

        # 파워링크 광고 영역
        ad_items = (
            soup.select('.ad_section li.lst') or
            soup.select('#power_link_body li.lst') or
            soup.select('[id^="pcPowerLink"] li.lst')
        )
        if not ad_items:
            ad_items = [li for li in soup.find_all('li', class_='lst')
                        if li.find('a', href=lambda h: h and 'ader.naver.com' in h)]

        for item in ad_items:
            info = self._parse_naver_ad(item, keyword)
            if info and self._add_business(info['domain'], info):
                count += 1
                print(f"      N광고✓ {info['name'][:20]} → {info['domain']}")

        # 일반 웹 검색 결과
        for a_tag in soup.select('a[href]'):
            href = a_tag.get('href', '')
            if not href.startswith('http') or 'naver.com' in href:
                if 'smartstore.naver.com' not in href:
                    continue

            try:
                parsed = urlparse(href)
                domain = parsed.netloc.lower().replace('www.', '')
            except Exception:
                continue

            if not domain or len(domain) < 4:
                continue

            domain_key = domain.split('/')[0]
            if 'smartstore.naver.com' in href:
                pp = parsed.path.strip('/').split('/')
                if pp and pp[0]:
                    domain_key = f"smartstore.naver.com/{pp[0]}"

            name = a_tag.get_text(strip=True)[:30] or domain_key.split('.')[0]
            if len(name) < 2:
                continue

            info = {
                'name': name, 'url': href, 'domain': domain_key,
                'phone': '', 'email': '', 'instagram': '',
                'desc': f'네이버: {keyword}', 'keyword': keyword,
                'tags': '', 'sublinks': '',
                'collected_at': datetime.now().isoformat(), 'source': 'naver',
            }
            if self._add_business(domain_key, info):
                count += 1
                if count <= 3:
                    print(f"      N✓ {name[:20]} → {domain_key}")

        # 네이버 검색 2페이지
        url2 = f"https://search.naver.com/search.naver?query={quote(keyword)}&where=web&start=16"
        html2 = self._http_get(url2, headers=headers)
        if html2:
            soup2 = BeautifulSoup(html2, 'html.parser')
            for a_tag in soup2.select('a[href]'):
                href = a_tag.get('href', '')
                if not href.startswith('http'):
                    continue
                try:
                    parsed = urlparse(href)
                    domain = parsed.netloc.lower().replace('www.', '')
                except Exception:
                    continue
                if not domain or len(domain) < 4:
                    continue
                domain_key = domain.split('/')[0]
                if 'smartstore.naver.com' in href:
                    pp = parsed.path.strip('/').split('/')
                    if pp and pp[0]:
                        domain_key = f"smartstore.naver.com/{pp[0]}"
                name = a_tag.get_text(strip=True)[:30] or domain_key.split('.')[0]
                if len(name) < 2:
                    continue
                info = {
                    'name': name, 'url': href, 'domain': domain_key,
                    'phone': '', 'email': '', 'instagram': '',
                    'desc': f'네이버: {keyword}', 'keyword': keyword,
                    'tags': '', 'sublinks': '',
                    'collected_at': datetime.now().isoformat(), 'source': 'naver',
                }
                if self._add_business(domain_key, info):
                    count += 1

        return count

    def _parse_naver_ad(self, item, keyword):
        """네이버 파워링크 광고 항목 파싱"""
        name = ''
        site_el = item.select_one('a.site')
        if site_el:
            name = site_el.get_text(strip=True)

        actual_domain = ''
        url_el = item.select_one('a.lnk_url')
        if url_el:
            actual_domain = url_el.get_text(strip=True).rstrip('/')

        if not actual_domain:
            url_area = item.select_one('.url_area')
            if url_area:
                for a in url_area.find_all('a'):
                    text = a.get_text(strip=True).rstrip('/')
                    if '.' in text and ' ' not in text and len(text) > 3:
                        actual_domain = text
                        break

        if not actual_domain:
            return None

        domain = actual_domain.lower().strip()
        if domain.startswith('http'):
            domain = urlparse(domain).netloc

        if 'smartstore' in domain:
            full_url = f"https://{actual_domain}"
            domain_key = actual_domain.lower().strip()
        else:
            domain_key = domain.split('/')[0]
            full_url = f"https://{domain_key}"

        if 'naver.com' in domain_key and 'smartstore' not in domain_key:
            return None

        titles = [t.get_text(strip=True) for t in item.select('span.lnk_tit') if t.get_text(strip=True)]
        desc_el = item.select_one('a.link_desc')
        desc = desc_el.get_text(strip=True) if desc_el else ''
        tags = [t.get_text(strip=True) for t in item.select('.keyword_area .item') if t.get_text(strip=True)]

        phone = ''
        etc_area = item.select_one('.etc_area .item')
        if etc_area:
            m = re.search(r'(0\d{1,2}[-.\s)]{1,3}\d{3,4}[-.\s]{1,2}\d{4})', etc_area.get_text())
            if m:
                phone = m.group(1)

        if not name or len(name) < 2:
            name = titles[0] if titles else domain_key.replace('www.', '').split('.')[0]

        return {
            'name': name[:30], 'url': full_url, 'domain': domain_key,
            'phone': phone, 'email': '', 'instagram': '',
            'desc': (desc or ' '.join(titles))[:100], 'keyword': keyword,
            'tags': ', '.join(tags[:5]), 'sublinks': '',
            'collected_at': datetime.now().isoformat(), 'source': 'naver_ad',
        }

    # ── [2] 구글 검색 (requests) ──
    def search_google(self, keyword):
        """구글 검색 — 여러 쿼리 변형 + 페이지네이션"""
        from bs4 import BeautifulSoup
        count = 0

        queries = [
            f"{keyword} 업체",
            f"{keyword} 업소용 구매",
            f"{keyword} 도매 판매",
        ]

        for query in queries:
            for start in [0, 10]:  # 2페이지까지
                try:
                    url = f"https://www.google.com/search?q={quote(query)}&num=20&hl=ko&start={start}"
                    html = self._http_get(url, timeout=10)
                    if not html:
                        continue

                    soup = BeautifulSoup(html, 'html.parser')
                    for a_tag in soup.select('a[href]'):
                        href = a_tag.get('href', '')
                        actual_url = ''
                        if '/url?q=' in href:
                            actual_url = href.split('/url?q=')[1].split('&')[0]
                        elif href.startswith('http') and 'google' not in href:
                            actual_url = href

                        if not actual_url:
                            continue

                        try:
                            parsed = urlparse(actual_url)
                            domain = parsed.netloc.lower().replace('www.', '')
                        except Exception:
                            continue

                        if not domain or len(domain) < 4:
                            continue

                        domain_key = domain.split('/')[0]
                        if 'smartstore.naver.com' in actual_url:
                            pp = parsed.path.strip('/').split('/')
                            if pp and pp[0]:
                                domain_key = f"smartstore.naver.com/{pp[0]}"

                        name = a_tag.get_text(strip=True)[:30] or domain_key.split('.')[0]
                        if len(name) < 2:
                            continue

                        info = {
                            'name': name, 'url': actual_url, 'domain': domain_key,
                            'phone': '', 'email': '', 'instagram': '',
                            'desc': f'구글: {keyword}', 'keyword': keyword,
                            'tags': '', 'sublinks': '',
                            'collected_at': datetime.now().isoformat(), 'source': 'google',
                        }
                        if self._add_business(domain_key, info):
                            count += 1
                            if count <= 3:
                                print(f"      G✓ {name[:20]} → {domain_key}")

                    time.sleep(random.uniform(2, 4))

                except Exception as e:
                    if '429' in str(e) or 'Too Many' in str(e):
                        print(f"      구글 제한 감지, 30초 대기...")
                        time.sleep(30)
                    continue

        return count

    # ── [3] 네이버 쇼핑 (requests) ──
    def search_naver_shopping(self, keyword):
        """네이버 쇼핑 검색 — 판매 업체(스마트스토어 등) 수집"""
        count = 0
        try:
            url = f"https://search.shopping.naver.com/search/all?query={quote(keyword)}&sort=rel"
            headers = {
                'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) '
                              'AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36',
                'Referer': 'https://shopping.naver.com/',
            }
            html = self._http_get(url, headers=headers)
            if not html:
                return count

            # JSON 데이터 추출 (네이버 쇼핑은 __NEXT_DATA__에 데이터를 넣음)
            nd_match = re.search(r'<script id="__NEXT_DATA__"[^>]*>(.+?)</script>', html)
            if nd_match:
                try:
                    nd = json.loads(nd_match.group(1))
                    products = (nd.get('props', {}).get('pageProps', {})
                                .get('initialState', {}).get('products', {})
                                .get('list', []))
                    for prod in products:
                        item = prod.get('item', {})
                        mall_name = item.get('mallName', '')
                        mall_url = item.get('mallProductUrl', '') or item.get('crUrl', '')
                        if not mall_name or not mall_url:
                            continue

                        try:
                            parsed = urlparse(mall_url)
                            domain = parsed.netloc.lower().replace('www.', '')
                        except Exception:
                            continue

                        domain_key = domain.split('/')[0]
                        if 'smartstore.naver.com' in mall_url:
                            pp = parsed.path.strip('/').split('/')
                            if pp and pp[0]:
                                domain_key = f"smartstore.naver.com/{pp[0]}"

                        info = {
                            'name': mall_name[:30], 'url': mall_url, 'domain': domain_key,
                            'phone': '', 'email': '', 'instagram': '',
                            'desc': f'쇼핑: {keyword}', 'keyword': keyword,
                            'tags': '', 'sublinks': '',
                            'collected_at': datetime.now().isoformat(), 'source': 'naver_shopping',
                        }
                        if self._add_business(domain_key, info):
                            count += 1
                except (json.JSONDecodeError, KeyError):
                    pass

            # HTML에서도 직접 추출 (fallback)
            store_urls = re.findall(r'https?://smartstore\.naver\.com/([a-zA-Z0-9_-]+)', html)
            for store_id in set(store_urls):
                domain_key = f"smartstore.naver.com/{store_id}"
                info = {
                    'name': store_id, 'url': f"https://smartstore.naver.com/{store_id}",
                    'domain': domain_key,
                    'phone': '', 'email': '', 'instagram': '',
                    'desc': f'쇼핑: {keyword}', 'keyword': keyword,
                    'tags': '', 'sublinks': '',
                    'collected_at': datetime.now().isoformat(), 'source': 'naver_shopping',
                }
                if self._add_business(domain_key, info):
                    count += 1

            # 일반 쇼핑몰 URL도 추출
            mall_urls = re.findall(r'"mallProductUrl"\s*:\s*"(https?://[^"]+)"', html)
            for murl in mall_urls:
                try:
                    parsed = urlparse(murl)
                    domain = parsed.netloc.lower().replace('www.', '')
                    domain_key = domain.split('/')[0]
                    if 'smartstore.naver.com' in murl:
                        pp = parsed.path.strip('/').split('/')
                        if pp and pp[0]:
                            domain_key = f"smartstore.naver.com/{pp[0]}"

                    mall_name = re.search(r'"mallName"\s*:\s*"([^"]+)"', html)
                    name = mall_name.group(1) if mall_name else domain_key.split('.')[0]

                    info = {
                        'name': name[:30], 'url': murl, 'domain': domain_key,
                        'phone': '', 'email': '', 'instagram': '',
                        'desc': f'쇼핑: {keyword}', 'keyword': keyword,
                        'tags': '', 'sublinks': '',
                        'collected_at': datetime.now().isoformat(), 'source': 'naver_shopping',
                    }
                    if self._add_business(domain_key, info):
                        count += 1
                except Exception:
                    continue

        except Exception as e:
            print(f"      쇼핑 오류: {e}")

        return count

    # ── [4] 네이버 지도/플레이스 (requests) ──
    def search_naver_map(self, keyword):
        """네이버 지도 검색 — 지역 업체 수집"""
        count = 0
        try:
            # 네이버 지도 검색 API
            url = f"https://map.naver.com/v5/api/search?caller=pcweb&query={quote(keyword)}&type=all&page=1&displayCount=20"
            headers = {
                'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) '
                              'AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36',
                'Referer': 'https://map.naver.com/',
            }
            resp_text = self._http_get(url, headers=headers)
            if not resp_text:
                return count

            try:
                data = json.loads(resp_text)
                places = data.get('result', {}).get('place', {}).get('list', [])
                for place in places:
                    name = place.get('name', '')
                    phone = place.get('tel', '') or place.get('phone', '')
                    address = place.get('address', '') or place.get('roadAddress', '')
                    homepage = place.get('homePage', '') or place.get('homepage', '')
                    category = place.get('category', '')
                    biz_num = place.get('businessNum', '')

                    if not name:
                        continue

                    # 도메인 키 생성
                    if homepage:
                        try:
                            parsed = urlparse(homepage if homepage.startswith('http') else f"https://{homepage}")
                            domain_key = parsed.netloc.lower().replace('www.', '').split('/')[0]
                        except Exception:
                            domain_key = name.replace(' ', '_')
                    else:
                        domain_key = f"map_{name.replace(' ', '_')}_{phone.replace('-', '')}"

                    if not domain_key or len(domain_key) < 3:
                        domain_key = f"map_{name[:20]}"

                    info = {
                        'name': name[:30],
                        'url': homepage or f"https://map.naver.com/v5/search/{quote(name)}",
                        'domain': domain_key,
                        'phone': phone, 'email': '', 'instagram': '',
                        'desc': f'지도: {category or keyword}', 'keyword': keyword,
                        'tags': category, 'sublinks': address[:50],
                        'collected_at': datetime.now().isoformat(), 'source': 'naver_map',
                    }
                    if self._add_business(domain_key, info):
                        count += 1
                        if count <= 2:
                            print(f"      M✓ {name[:20]} ({phone})")

            except json.JSONDecodeError:
                pass

            # 2페이지
            url2 = f"https://map.naver.com/v5/api/search?caller=pcweb&query={quote(keyword)}&type=all&page=2&displayCount=20"
            resp2 = self._http_get(url2, headers=headers)
            if resp2:
                try:
                    data2 = json.loads(resp2)
                    places2 = data2.get('result', {}).get('place', {}).get('list', [])
                    for place in places2:
                        name = place.get('name', '')
                        phone = place.get('tel', '') or place.get('phone', '')
                        homepage = place.get('homePage', '') or place.get('homepage', '')
                        category = place.get('category', '')
                        if not name:
                            continue
                        if homepage:
                            try:
                                parsed = urlparse(homepage if homepage.startswith('http') else f"https://{homepage}")
                                domain_key = parsed.netloc.lower().replace('www.', '').split('/')[0]
                            except Exception:
                                domain_key = f"map_{name.replace(' ', '_')}"
                        else:
                            domain_key = f"map_{name.replace(' ', '_')}_{phone.replace('-', '')}"
                        info = {
                            'name': name[:30],
                            'url': homepage or f"https://map.naver.com/v5/search/{quote(name)}",
                            'domain': domain_key,
                            'phone': phone, 'email': '', 'instagram': '',
                            'desc': f'지도: {category or keyword}', 'keyword': keyword,
                            'tags': category, 'sublinks': '',
                            'collected_at': datetime.now().isoformat(), 'source': 'naver_map',
                        }
                        if self._add_business(domain_key, info):
                            count += 1
                except (json.JSONDecodeError, KeyError):
                    pass

        except Exception as e:
            print(f"      지도 오류: {e}")

        return count

    # ── 스마트스토어 API ──
    def _try_smartstore_api(self, store_url):
        result = {'email': '', 'name': '', 'phone': '', 'desc': ''}
        try:
            parsed = urlparse(store_url)
            pp = parsed.path.strip('/').split('/')
            if not pp or not pp[0]:
                return result
            store_id = pp[0]
            http = self._get_http()

            for api_url in [
                f"https://smartstore.naver.com/i/v1/stores/{store_id}/seller-info",
                f"https://smartstore.naver.com/{store_id}/profile",
            ]:
                try:
                    resp = http.get(api_url, timeout=10, headers={'Referer': store_url})
                    if resp.status_code == 200:
                        emails = re.findall(r'[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}', resp.text)
                        valid = [e.lower() for e in emails
                                 if 'naver.com' not in e.lower().split('@')[1]
                                 and 'navercorp' not in e.lower()]
                        if valid:
                            result['email'] = valid[0]
                            return result
                except Exception:
                    continue

            # __NEXT_DATA__에서 추출
            try:
                resp = http.get(store_url, timeout=10)
                if resp.status_code == 200:
                    nd = re.search(r'<script id="__NEXT_DATA__"[^>]*>(.+?)</script>', resp.text)
                    if nd:
                        emails = re.findall(r'[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}', nd.group(1))
                        valid = [e.lower() for e in emails if 'naver.com' not in e.lower().split('@')[1]]
                        if valid:
                            result['email'] = valid[0]
                    nm = re.search(r'"storeName"\s*:\s*"([^"]+)"', resp.text)
                    if nm:
                        result['name'] = nm.group(1)
            except Exception:
                pass

            # 판매자 정보 페이지 직접 요청
            seller_urls = [
                f"https://smartstore.naver.com/i/v1/stores/{store_id}/seller",
                f"https://smartstore.naver.com/{store_id}/seller-info",
            ]
            for surl in seller_urls:
                try:
                    resp = http.get(surl, timeout=10, headers={'Referer': store_url})
                    if resp.status_code == 200:
                        emails = re.findall(r'[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}', resp.text)
                        valid = [e.lower() for e in emails
                                 if 'naver.com' not in e.lower().split('@')[1]
                                 and 'navercorp' not in e.lower()]
                        if valid:
                            result['email'] = valid[0]
                            phones = re.findall(r'(0\d{1,2}[-.\s)]{1,3}\d{3,4}[-.\s]{1,2}\d{4})', resp.text)
                            if phones:
                                result['phone'] = phones[0]
                            return result
                except Exception:
                    continue

        except Exception:
            pass
        return result

    # ── 이메일 추출 유틸 ──
    def _extract_emails_from_html(self, html_source):
        import html as html_mod
        emails = []
        emails.extend(re.findall(r'[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}', html_source))
        emails.extend(re.findall(r'mailto:([a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,})', html_source))
        for user, domain in re.findall(
            r'([a-zA-Z0-9._%+\-]+)\s*[\[(\{]\s*(?:at|@|골뱅이)\s*[\])\}]\s*([a-zA-Z0-9.\-]+\.[a-zA-Z]{2,})',
            html_source, re.IGNORECASE
        ):
            emails.append(f"{user}@{domain}")
        decoded = html_mod.unescape(html_source)
        if decoded != html_source:
            emails.extend(re.findall(r'[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}', decoded))
        emails.extend(re.findall(r'"email"\s*:\s*"([^"]+@[^"]+)"', html_source))
        emails.extend(re.findall(r'content="([^"]*@[^"]*\.[a-zA-Z]{2,})"', html_source))

        skip = ['example.com', 'test.com', 'domain.com', 'email.com',
                'yoursite', 'mysite', 'sentry.io', 'wixpress', 'w3.org',
                'googleapis', 'cloudflare', 'schema.org', 'jquery',
                '.png', '.jpg', '.gif', '.js', '.css', '.svg',
                'noreply', 'no-reply', 'mailer-daemon', 'postmaster',
                'webpack', 'babel', 'eslint', 'prettier']
        valid = set()
        for e in emails:
            e = e.lower().strip().rstrip('.')
            if any(s in e for s in skip) or len(e) < 6 or '.' not in e.split('@')[-1]:
                continue
            if re.match(r'^[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}$', e):
                valid.add(e)
        return valid

    def _pick_best_email(self, valid_emails):
        if not valid_emails:
            return ''
        for prefix in ['info@', 'contact@', 'admin@', 'sales@', 'cs@', 'help@',
                        'master@', 'ceo@', 'biz@', 'service@', 'support@',
                        'marketing@', 'pr@', 'manager@', 'office@']:
            for e in valid_emails:
                if e.startswith(prefix):
                    return e
        return min(valid_emails, key=len)

    def _extract_phone_from_text(self, text):
        for pat in [r'(0\d{1,2}[-.\s)]{1,3}\d{3,4}[-.\s]{1,2}\d{4})',
                     r'(1[0-9]{3}[-.\s]\d{4})', r'(02[-.\s]\d{3,4}[-.\s]\d{4})']:
            m = re.search(pat, text)
            if m:
                return m.group(1).strip()
        return ''

    # ── 일반 웹사이트 이메일 추출 (requests) ──
    def extract_email_from_website(self, domain, url):
        from bs4 import BeautifulSoup
        result = {'email': '', 'name': '', 'phone': '', 'desc': ''}

        if 'smartstore.naver.com' in domain or 'smartstore.naver.com' in url:
            return self._try_smartstore_api(url)

        # 지도 수집 업체는 홈페이지가 없을 수 있음
        if domain.startswith('map_'):
            return result

        base = url if url.startswith('http') else f"https://{domain}"
        visited = set()
        discovered = []

        def fetch(page_url):
            nonlocal result
            if page_url in visited:
                return False
            visited.add(page_url)
            html = self._http_get(page_url)
            if not html:
                return False

            if not result['name']:
                m = re.search(r'<title>([^<]+)</title>', html, re.IGNORECASE)
                if m:
                    t = m.group(1).strip()
                    for sep in [' - ', ' | ', ' :: ', ' – ', ' — ']:
                        if sep in t:
                            t = t.split(sep)[0].strip()
                    if 2 <= len(t) <= 30:
                        result['name'] = t

            valid = self._extract_emails_from_html(html)
            if valid and not result['email']:
                result['email'] = self._pick_best_email(valid)
            if not result['phone']:
                result['phone'] = self._extract_phone_from_text(html)
            if result['email']:
                return True

            if page_url == base and not discovered:
                soup = BeautifulSoup(html, 'html.parser')
                kws = ['회사소개', '기업소개', '연락처', '문의', 'about',
                       'company', 'contact', '개인정보', '사업자정보']
                for a in soup.find_all('a', href=True):
                    txt = a.get_text(strip=True).lower()
                    hr = a['href'].lower()
                    if any(k in txt or k in hr for k in kws):
                        full = urljoin(base, a['href'])
                        if full.startswith('http') and full not in visited:
                            ld = urlparse(full).netloc.lower()
                            if domain in ld or ld in domain:
                                discovered.append(full)
                                if len(discovered) >= 3:
                                    break
            return False

        if fetch(base):
            return result
        for p in discovered[:3]:
            if fetch(p):
                return result
        for path in ['/company', '/about', '/contact']:
            if fetch(urljoin(base, path)):
                return result
        return result

    # ── 메인 수집 루프 ──
    def collect_all(self):
        print("\n" + "=" * 60)
        print(" 마트마트 광고주 자동 수집 시작 (브라우저 안 씀)")
        print("=" * 60)

        remaining = [k for k in KEYWORDS if k not in self.done_keywords]
        remaining_shop = [k for k in SHOPPING_KEYWORDS if k not in self.done_shopping]
        remaining_map = [k for k in MAP_KEYWORDS if k not in self.done_map]
        print(f"\n총 키워드: 네이버/구글 {len(KEYWORDS)}개 + 쇼핑 {len(SHOPPING_KEYWORDS)}개 + 지도 {len(MAP_KEYWORDS)}개")
        print(f"완료: 네이버/구글 {len(self.done_keywords)}개 + 쇼핑 {len(self.done_shopping)}개 + 지도 {len(self.done_map)}개")
        print(f"현재 수집: {len(self.collected)}개\n")

        try:
            # [1/5] 네이버 검색
            if remaining:
                print("─" * 40)
                print(f" [1/5] 네이버 검색 ({len(remaining)}개 키워드)")
                print("─" * 40)
                for i, kw in enumerate(remaining, 1):
                    print(f"\n  ({i}/{len(remaining)}) '{kw}'")
                    n = self.search_naver(kw)
                    print(f"    → +{n}개 (누적 {len(self.collected)})")
                    self.done_keywords.add(kw)
                    if i % 5 == 0:
                        self.save_progress()
                    time.sleep(random.uniform(2, 4))
                self.save_progress()
                print(f"\n  네이버 완료! 누적 {len(self.collected)}개")
            else:
                print("  네이버 키워드 전부 완료 → 스킵")

            # [2/5] 구글 검색
            print("\n" + "─" * 40)
            print(f" [2/5] 구글 검색 ({len(KEYWORDS)}개 키워드)")
            print("─" * 40)
            before = len(self.collected)
            for i, kw in enumerate(KEYWORDS, 1):
                print(f"\n  ({i}/{len(KEYWORDS)}) 구글: '{kw}'")
                self.search_google(kw)
                if i % 5 == 0:
                    self.save_progress()
                    time.sleep(random.uniform(3, 6))
                else:
                    time.sleep(random.uniform(1, 3))
            print(f"\n  구글 +{len(self.collected) - before}개 (총 {len(self.collected)})")
            self.save_progress()

            # [3/5] 네이버 쇼핑
            if remaining_shop:
                print("\n" + "─" * 40)
                print(f" [3/5] 네이버 쇼핑 ({len(remaining_shop)}개 키워드)")
                print("─" * 40)
                before = len(self.collected)
                for i, kw in enumerate(remaining_shop, 1):
                    print(f"\n  ({i}/{len(remaining_shop)}) 쇼핑: '{kw}'", end="")
                    n = self.search_naver_shopping(kw)
                    print(f" → +{n}개")
                    self.done_shopping.add(kw)
                    if i % 5 == 0:
                        self.save_progress()
                    time.sleep(random.uniform(1, 3))
                print(f"\n  쇼핑 +{len(self.collected) - before}개 (총 {len(self.collected)})")
                self.save_progress()
            else:
                print("  쇼핑 키워드 전부 완료 → 스킵")

            # [4/5] 네이버 지도
            if remaining_map:
                print("\n" + "─" * 40)
                print(f" [4/5] 네이버 지도 ({len(remaining_map)}개 키워드)")
                print("─" * 40)
                before = len(self.collected)
                for i, kw in enumerate(remaining_map, 1):
                    print(f"\n  ({i}/{len(remaining_map)}) 지도: '{kw}'", end="")
                    n = self.search_naver_map(kw)
                    print(f" → +{n}개")
                    self.done_map.add(kw)
                    time.sleep(random.uniform(1, 3))
                print(f"\n  지도 +{len(self.collected) - before}개 (총 {len(self.collected)})")
                self.save_progress()
            else:
                print("  지도 키워드 전부 완료 → 스킵")

            # [5/5] 이메일 추출
            print("\n" + "─" * 40)
            print(" [5/5] 이메일 추출 (브라우저 안 씀)")
            print("─" * 40)
            no_email = [d for d, i in self.collected.items() if not i.get('email')]
            normals = [d for d in no_email if 'smartstore' not in d and not d.startswith('map_')]
            smarts = [d for d in no_email if 'smartstore' in d]
            maps = [d for d in no_email if d.startswith('map_') and self.collected[d].get('url', '').startswith('http')]
            print(f"\n  미추출: {len(no_email)}개 (일반 {len(normals)} + 스마트스토어 {len(smarts)} + 지도(홈페이지) {len(maps)})")

            # 일반 사이트
            if normals:
                print(f"\n  --- 일반 사이트 ({len(normals)}개) ---")
                for i, d in enumerate(normals, 1):
                    info = self.collected[d]
                    print(f"  ({i}/{len(normals)}) {info['name'][:15]}... ", end="", flush=True)
                    r = self.extract_email_from_website(d, info['url'])
                    if r.get('email'):
                        self.collected[d]['email'] = r['email']
                        if r.get('name') and len(r['name']) > len(info.get('name', '')):
                            self.collected[d]['name'] = r['name']
                        if r.get('phone') and not info.get('phone'):
                            self.collected[d]['phone'] = r['phone']
                        print(f"✓ {r['email']}")
                    else:
                        print("✗")
                    if i % 20 == 0:
                        self.save_progress()
                    time.sleep(random.uniform(0.5, 1.5))
            self.save_progress()

            # 스마트스토어
            if smarts:
                print(f"\n  --- 스마트스토어 API ({len(smarts)}개) ---")
                for i, d in enumerate(smarts, 1):
                    info = self.collected[d]
                    print(f"  ({i}/{len(smarts)}) {info['name'][:15]}... ", end="", flush=True)
                    api = self._try_smartstore_api(info['url'])
                    if api.get('email'):
                        self.collected[d]['email'] = api['email']
                        if api.get('name'):
                            self.collected[d]['name'] = api['name']
                        print(f"✓ {api['email']}")
                    else:
                        print("✗")
                    time.sleep(random.uniform(0.5, 1))
            self.save_progress()

            # 지도에서 수집한 업체 중 홈페이지 있는 곳
            maps_with_homepage = [d for d in maps
                                  if not self.collected[d]['url'].startswith('https://map.naver.com')]
            if maps_with_homepage:
                print(f"\n  --- 지도 업체 홈페이지 ({len(maps_with_homepage)}개) ---")
                for i, d in enumerate(maps_with_homepage, 1):
                    info = self.collected[d]
                    print(f"  ({i}/{len(maps_with_homepage)}) {info['name'][:15]}... ", end="", flush=True)
                    r = self.extract_email_from_website(d, info['url'])
                    if r.get('email'):
                        self.collected[d]['email'] = r['email']
                        print(f"✓ {r['email']}")
                    else:
                        print("✗")
                    if i % 20 == 0:
                        self.save_progress()
                    time.sleep(random.uniform(0.5, 1.5))
            self.save_progress()

        except KeyboardInterrupt:
            print("\n\n  중단됨! 저장 중...")
            self.save_progress()

        we = sum(1 for v in self.collected.values() if v.get('email'))
        print(f"\n{'=' * 60}")
        print(f" 수집 결과: 총 {len(self.collected)}개 / 이메일 {we}개")
        print(f"{'=' * 60}")
        self.save_to_excel()
        return we

    # ── 엑셀 저장 ──
    def save_to_excel(self):
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        wb = Workbook()
        ws = wb.active
        ws.title = "광고주 리스트"

        hfill = PatternFill(start_color="1a73e8", end_color="1a73e8", fill_type="solid")
        hfont = Font(name="맑은 고딕", size=11, bold=True, color="FFFFFF")
        tb = Border(left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin'))

        headers = ["No.", "카테고리", "업체명", "이메일", "인스타그램", "전화번호", "웹사이트", "업체설명", "출처", "발송상태"]
        widths = [6, 18, 22, 28, 22, 16, 32, 35, 12, 10]

        for col, (h, w) in enumerate(zip(headers, widths), 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.fill = hfill
            cell.font = hfont
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = tb
            cl = chr(64 + col) if col <= 26 else 'A' + chr(64 + col - 26)
            ws.column_dimensions[cl].width = w

        sorted_data = sorted(self.collected.values(),
                             key=lambda x: (0 if x.get('email') else 1, x.get('keyword', '')))

        for i, info in enumerate(sorted_data, 1):
            vals = [
                i, info.get('keyword', ''), info.get('name', ''),
                info.get('email', ''), info.get('instagram', ''),
                info.get('phone', ''), info.get('url', ''),
                info.get('desc', ''), info.get('source', ''),
                '미발송'
            ]
            for col, val in enumerate(vals, 1):
                cell = ws.cell(row=i + 1, column=col, value=val)
                cell.border = tb
                cell.alignment = Alignment(vertical='center')
                if col == 4 and val:
                    cell.font = Font(color="0000FF")
                if col == 5 and val:
                    cell.font = Font(color="8B00FF")

        wb.save(EXCEL_FILE)
        print(f"\n  엑셀 저장: {EXCEL_FILE}")


# ============================================================
#  2단계: 인스타그램 업체 수집 & DM 발송
# ============================================================
class InstagramManager:
    """인스타그램 비즈니스 계정 수집 + DM 발송 (instagrapi — 브라우저 안 씀)"""

    def __init__(self):
        self.client = None
        self.dm_sent = {}
        self.dm_today_count = 0
        self._load_dm_progress()

    def _load_dm_progress(self):
        if os.path.exists(INSTA_DM_PROGRESS_FILE):
            try:
                with open(INSTA_DM_PROGRESS_FILE, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                    self.dm_sent = data.get('sent', {})
                    today = datetime.now().strftime('%Y-%m-%d')
                    self.dm_today_count = sum(
                        1 for v in self.dm_sent.values()
                        if v.startswith(today)
                    )
                    print(f"  DM 진행상황: 총 {len(self.dm_sent)}개 발송, 오늘 {self.dm_today_count}개")
            except Exception:
                self.dm_sent = {}
                self.dm_today_count = 0

    def _save_dm_progress(self):
        with open(INSTA_DM_PROGRESS_FILE, 'w', encoding='utf-8') as f:
            json.dump({
                'sent': self.dm_sent,
                'last_update': datetime.now().isoformat(),
            }, f, ensure_ascii=False, indent=2)

    def _challenge_code_handler(self, username, choice):
        print(f"\n  인스타그램 보안 인증이 필요합니다!")
        print(f"  인증 방법: {choice}")
        print(f"  → 인스타 앱 또는 이메일/SMS에서 인증 코드를 확인하세요.")
        code = input("  인증 코드 입력: ").strip()
        return code

    def login(self):
        import getpass

        try:
            from instagrapi import Client
        except ImportError:
            print("  instagrapi가 설치되지 않았습니다.")
            print("  → pip3 install instagrapi")
            return False

        self.client = Client()
        self.client.delay_range = [2, 5]
        self.client.challenge_code_handler = self._challenge_code_handler

        # 저장된 세션이 있으면 먼저 시도
        if os.path.exists(INSTA_SESSION_FILE):
            try:
                self.client.load_settings(INSTA_SESSION_FILE)
                self.client.login(INSTA_USERNAME, INSTA_PASSWORD)
                self.client.get_timeline_feed()
                print("  인스타그램 로그인 성공 (세션 재사용)")
                return True
            except Exception:
                print("  세션 만료, 새로 로그인...")
                try:
                    os.remove(INSTA_SESSION_FILE)
                except Exception:
                    pass
                self.client = Client()
                self.client.delay_range = [2, 5]
                self.client.challenge_code_handler = self._challenge_code_handler

        # 비밀번호: 설정된 값 사용, 없으면 터미널 입력
        if INSTA_PASSWORD:
            password = INSTA_PASSWORD
        else:
            print(f"\n  인스타그램 계정: {INSTA_USERNAME}")
            password = getpass.getpass("  비밀번호 입력: ").strip()
            if not password:
                print("  비밀번호를 입력하지 않았습니다.")
                return False

        try:
            print("  로그인 시도 중...")
            self.client.login(INSTA_USERNAME, password)
            self.client.dump_settings(INSTA_SESSION_FILE)
            print("  인스타그램 로그인 성공! (세션 저장됨 → 다음엔 자동 로그인)")
            return True
        except Exception as e:
            error_str = str(e).lower()
            if 'challenge' in error_str or 'checkpoint' in error_str:
                print(f"\n  인스타그램 보안 인증 필요")
                print("  해결 방법:")
                print("    1. 인스타 앱 → '네, 접니다' 승인")
                print("    2. 승인 후 이 스크립트를 다시 실행하세요")
            elif 'password' in error_str or 'bad_password' in error_str:
                print(f"\n  비밀번호가 틀렸습니다. 다시 확인해 주세요.")
            elif 'ip' in error_str or 'blacklist' in error_str:
                print(f"\n  IP가 차단되었습니다.")
                print("  해결 방법:")
                print("    1. 먼저 같은 기기에서 인스타 앱/브라우저로 로그인")
                print("    2. '비정상 로그인 시도' 알림이 있다면 승인")
                print("    3. 잠시 후 (5~10분) 이 스크립트를 다시 실행")
            else:
                print(f"  로그인 실패: {e}")
            self.client = None
            return False

    def collect_business_accounts(self):
        if not self.client:
            print("  인스타그램 미로그인 → 수집 스킵")
            return {}

        accounts = {}

        # 1. 해시태그 기반
        print(f"\n  --- 해시태그 검색 ({len(INSTA_HASHTAGS)}개) ---")
        for i, tag in enumerate(INSTA_HASHTAGS, 1):
            print(f"  ({i}/{len(INSTA_HASHTAGS)}) #{tag}... ", end="", flush=True)
            try:
                medias = self.client.hashtag_medias_top(tag, amount=9)
                found = 0
                for media in medias:
                    try:
                        user = self.client.user_info(media.user.pk)
                        if self._is_business_target(user):
                            username = user.username
                            if username not in accounts and username != INSTA_USERNAME:
                                accounts[username] = {
                                    'username': username,
                                    'full_name': user.full_name or '',
                                    'bio': (user.biography or '')[:100],
                                    'follower_count': user.follower_count,
                                    'is_business': user.is_business,
                                    'category': user.business_category_name or '',
                                    'contact_phone': user.business_phone_number or '',
                                    'contact_email': user.business_contact_method or '',
                                    'external_url': user.external_url or '',
                                    'user_id': str(user.pk),
                                    'hashtag': tag,
                                }
                                found += 1
                    except Exception:
                        continue
                print(f"{found}개 수집")
                time.sleep(random.uniform(3, 6))
            except Exception as e:
                print(f"오류: {e}")
                time.sleep(random.uniform(5, 10))

        # 2. 키워드 검색 기반
        print(f"\n  --- 키워드 검색 ({len(INSTA_SEARCH_KEYWORDS)}개) ---")
        for i, kw in enumerate(INSTA_SEARCH_KEYWORDS, 1):
            print(f"  ({i}/{len(INSTA_SEARCH_KEYWORDS)}) '{kw}'... ", end="", flush=True)
            try:
                users = self.client.search_users_v1(kw, count=20)
                found = 0
                for u in users:
                    try:
                        user = self.client.user_info(u.pk)
                        if self._is_business_target(user):
                            username = user.username
                            if username not in accounts and username != INSTA_USERNAME:
                                accounts[username] = {
                                    'username': username,
                                    'full_name': user.full_name or '',
                                    'bio': (user.biography or '')[:100],
                                    'follower_count': user.follower_count,
                                    'is_business': user.is_business,
                                    'category': user.business_category_name or '',
                                    'contact_phone': user.business_phone_number or '',
                                    'contact_email': user.business_contact_method or '',
                                    'external_url': user.external_url or '',
                                    'user_id': str(user.pk),
                                    'hashtag': f'검색:{kw}',
                                }
                                found += 1
                    except Exception:
                        continue
                print(f"{found}개 수집")
                time.sleep(random.uniform(3, 6))
            except Exception as e:
                print(f"오류: {e}")
                time.sleep(random.uniform(5, 10))

        print(f"\n  인스타 수집 완료: {len(accounts)}개 비즈니스 계정")
        return accounts

    def _is_business_target(self, user):
        if user.follower_count < 100 or user.follower_count > 1000000:
            return False
        if user.is_private:
            return False
        if user.is_business:
            return True
        bio = (user.biography or '').lower()
        biz_keywords = [
            '업소용', '마트', '냉장고', '냉동', '쇼케이스', '진열대',
            '인테리어', '설비', 'pos', '키오스크', '도매', '납품',
            '공급', '제조', '시공', '설치', '유통', '배송',
            '냉동탑차', '식자재', '포장', '전자저울', '간판',
        ]
        return any(kw in bio for kw in biz_keywords)

    def send_dm(self, username, user_id, company_name=''):
        if not self.client:
            return False
        if username in self.dm_sent:
            print(f"이미발송 ", end="", flush=True)
            return False
        if self.dm_today_count >= DAILY_DM_LIMIT:
            print(f"일일한도 ", end="", flush=True)
            return False
        try:
            name = company_name or username
            template = random.choice(DM_TEMPLATES)
            message = template.format(company=name)
            self.client.direct_send(message, [int(user_id)])
            self.dm_sent[username] = datetime.now().isoformat()
            self.dm_today_count += 1
            self._save_dm_progress()
            return True
        except Exception as e:
            error_str = str(e).lower()
            if 'challenge' in error_str or 'checkpoint' in error_str:
                print(f"인증필요 ", end="", flush=True)
            elif 'feedback_required' in error_str or 'spam' in error_str:
                print(f"스팸감지! 즉시 중단합니다.", flush=True)
                self.dm_today_count = DAILY_DM_LIMIT
            else:
                print(f"오류({e}) ", end="", flush=True)
            return False

    def send_dms_batch(self, accounts, limit=None):
        if not self.client:
            print("  인스타그램 미로그인")
            return 0
        targets = {k: v for k, v in accounts.items() if k not in self.dm_sent}
        if limit:
            targets = dict(list(targets.items())[:limit])
        remaining_today = DAILY_DM_LIMIT - self.dm_today_count
        if remaining_today <= 0:
            print(f"  오늘 DM 한도({DAILY_DM_LIMIT}개) 소진됨")
            return 0
        actual_targets = dict(list(targets.items())[:remaining_today])
        print(f"\n  DM 발송 대상: {len(actual_targets)}개 (일일 잔여: {remaining_today})")
        success = 0
        for i, (username, info) in enumerate(actual_targets.items(), 1):
            if self.dm_today_count >= DAILY_DM_LIMIT:
                print(f"\n  일일 한도 도달!")
                break
            name = info.get('full_name', '') or info.get('username', '')
            print(f"  [{i}/{len(actual_targets)}] @{username} ({name})... ", end="", flush=True)
            if self.send_dm(username, info['user_id'], name):
                print("✓")
                success += 1
            else:
                print("✗")
            if i < len(actual_targets) and self.dm_today_count < DAILY_DM_LIMIT:
                delay = random.uniform(*DELAY_BETWEEN_DMS)
                print(f"         {delay:.0f}초 대기...", flush=True)
                time.sleep(delay)
        print(f"\n  DM 발송 완료: 성공 {success}개 / 오늘 총 {self.dm_today_count}개")
        return success


# ============================================================
#  3단계: 이메일 발송
# ============================================================
def make_email_html(company_name, category):
    return f"""<!DOCTYPE html>
<html><head><meta charset="utf-8">
<style>
  body {{
    font-family: 'Apple SD Gothic Neo', 'Malgun Gothic', Arial, sans-serif;
    font-size: 15px;
    line-height: 2.0;
    color: #1a1a1a;
    background: #ffffff;
    margin: 0; padding: 0;
  }}
  .wrap {{
    max-width: 580px;
    margin: 40px auto;
    padding: 0 20px;
  }}
  p {{ margin: 0 0 18px 0; }}
  ul {{ margin: 0 0 18px 0; padding-left: 20px; }}
  ul li {{ margin-bottom: 6px; }}
  hr {{ border: none; border-top: 1px solid #ddd; margin: 30px 0; }}
  .sig {{ font-size: 14px; color: #444; line-height: 1.9; }}
  .unsub {{ font-size: 12px; color: #aaa; margin-top: 14px; }}
</style>
</head>
<body>
<div class="wrap">

  <p><b>{company_name}</b> 담당자님께,</p>

  <p>안녕하세요. 네이버 카페 <b>마트마트</b>를 운영하고 있는 김우진 대표입니다.</p>

  <p>
    저희 카페가 마트·슈퍼마켓 업계 종사자들을 위한 커뮤니티이다 보니,
    귀사의 제품과 서비스가 저희 회원분들께 실질적인 도움이 될 수 있겠다는 판단 아래 연락 드리게 되었습니다.
  </p>

  <p>
    마트마트는 전국 마트·슈퍼마켓 운영자, 정육·수산 종사자 등 실제 매장을 운영하는 업계 종사자
    <b>22,000명 이상</b>이 모인 폐쇄형 전문 커뮤니티입니다.
    가입 승인제로 운영되어 일반 소비자 없이 업계 종사자만 활동하며,
    최근 30일 기준 월간 조회수 393,723회, 회원의 83.5%가 30~60대 실질 결정권자, 남성 비율 90.5%입니다.
  </p>

  <p>저희가 제공하는 광고 상품은 아래와 같습니다.</p>

  <ul>
    <li>PC 메인 대문 배너 광고</li>
    <li>모바일 게시글 하단 배너 광고</li>
    <li>파트너사 전용 홍보 게시판 개설 및 운영</li>
    <li>마트마트 공식 파트너 엠블럼 부여 (카페 매니저 등급)</li>
    <li>공동구매 프로모션 (월 1회, 협의 진행)</li>
    <li>전체 회원 대상 이메일 발송 (월 1회, Premium 상품)</li>
  </ul>

  <p>
    각 카테고리당 1개 업체만 독점 입점 가능하며, 자세한 내용은 첨부된 제안서에 정리되어 있습니다.
    부담 없이 살펴보시고 궁금한 점은 언제든 회신 주세요.
  </p>

  <p>감사합니다.</p>

  <hr>

  <div class="sig">
    <b>김우진</b> 대표<br>
    마트마트 네이버 카페<br>
    이메일: gimuuuujin@gmail.com<br>
    카페: https://cafe.naver.com/martmart
  </div>

  <p class="unsub">본 메일은 귀사의 공개된 연락처를 통해 발송되었습니다. 수신을 원치 않으시면 회신해 주세요.</p>

</div>
</body></html>"""


# 발송 차단 이메일 목록 (중복 발송 방지 / 수동 제외)
EMAIL_BLOCKLIST = {
    'info@nfctagfactory.com',
}


def load_prospects_from_excel():
    from openpyxl import load_workbook
    wb = load_workbook(EXCEL_FILE, data_only=True)
    ws = wb['광고주 리스트']
    prospects = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row and len(row) >= 10:
            no, cat, company, email, insta, phone, web, desc, source, status = row[:10]
            email_str = str(email).strip().lower()
            if (email and '@' in email_str
                    and str(status) != '발송완료'
                    and email_str not in EMAIL_BLOCKLIST):
                prospects.append({
                    'company': str(company), 'category': str(cat),
                    'email': str(email), 'row_num': no,
                })
    return prospects


DASHBOARD_FILE = os.path.join(_BASE_DIR, "마트마트_통합발송_검토.xlsx")

def _update_dashboard(email_addr):
    """통합발송_검토.xlsx 이메일 시트 + 대시보드 동기화 (단건)"""
    try:
        from openpyxl import load_workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        wb = load_workbook(DASHBOARD_FILE)
        ws_email = wb['📧 이메일 발송 대상']
        ws_db    = wb['📊 대시보드']

        # 이메일 시트: 해당 이메일 찾아서 발송완료 표시
        found = False
        for row in ws_email.iter_rows(min_row=2):
            if row[3].value and str(row[3].value).strip().lower() == email_addr.lower():
                row[1].value = '발송완료'
                row[1].fill = PatternFill(start_color='D4EDDA', end_color='D4EDDA', fill_type='solid')
                row[1].font = Font(color='155724', bold=True, size=10)
                row[1].alignment = Alignment(horizontal='center', vertical='center')
                found = True
                break

        if found:
            # 대시보드: 완료/대기 건수 재계산
            rows = [r for r in ws_email.iter_rows(min_row=2, values_only=True) if r and r[0]]
            total = len(rows)
            done  = sum(1 for r in rows if r[1] == '발송완료')
            pending = total - done
            ws_db.cell(row=4, column=1).value = total
            ws_db.cell(row=5, column=1).value = f'완료 {done}건 / 대기 {pending}건'
            wb.save(DASHBOARD_FILE)
    except Exception as e:
        print(f'  ⚠ 대시보드 업데이트 실패: {e}')


def mark_sent_in_excel(row_num, email_addr=None, _wb_cache={}):
    """발송 완료 즉시 엑셀에 반영 — 워크북을 메모리에 유지해서 파일 잠금 최소화"""
    try:
        from openpyxl import load_workbook
        # 워크북을 한 번만 열고 메모리에 유지
        if 'wb' not in _wb_cache:
            _wb_cache['wb'] = load_workbook(EXCEL_FILE)
        wb = _wb_cache['wb']
        ws = wb['광고주 리스트']
        found = False
        for row in ws.iter_rows(min_row=2):
            if row[0].value == row_num:
                row[9].value = '발송완료'
                found = True
                break
        if not found:
            print(f"  ⚠ row_num={row_num} 못 찾음")
            return
        # 저장 재시도 (Excel이 열려있어도 최대 3번 시도)
        for attempt in range(3):
            try:
                wb.save(EXCEL_FILE)
                print(f"     → 엑셀 ✓ (No.{row_num})")
                if email_addr:
                    _update_dashboard(email_addr)
                return
            except Exception:
                if attempt < 2:
                    import time as _t; _t.sleep(1)
        print(f"  ⚠ 엑셀 저장 실패 (No.{row_num}) — Excel 파일 닫아주세요")
    except Exception as e:
        print(f"  ⚠ 엑셀 업데이트 실패: {e}")


PROPOSAL_PDF = os.path.join(os.path.dirname(os.path.abspath(__file__)), '마트마트카페_광고제안서.pdf')


def _attach_pdf(msg):
    """제안서 PDF를 이메일에 첨부. 파일 없으면 경고만 출력."""
    if not os.path.exists(PROPOSAL_PDF):
        print(f"  ⚠ 제안서 PDF 없음 (건너뜀): {PROPOSAL_PDF}")
        return
    with open(PROPOSAL_PDF, 'rb') as f:
        part = MIMEBase('application', 'octet-stream')
        part.set_payload(f.read())
    encoders.encode_base64(part)
    part.add_header('Content-Disposition', 'attachment', filename='마트마트카페_광고제안서.pdf')
    msg.attach(part)


def _build_msg(to_addr, subject, company, category):
    """HTML 본문 + PDF 첨부 이메일 메시지 생성."""
    msg = MIMEMultipart('mixed')
    msg['From'] = f"김우진 대표 | 마트마트 <{GMAIL_EMAIL}>"
    msg['To'] = to_addr
    msg['Subject'] = subject
    # HTML 본문을 'related' 파트로 감싸서 첨부와 분리
    alt = MIMEMultipart('alternative')
    alt.attach(MIMEText(make_email_html(company, category), 'html', 'utf-8'))
    msg.attach(alt)
    _attach_pdf(msg)
    return msg


def load_email_daily_count() -> int:
    """오늘 날짜 기준 이메일 발송 건수 로드"""
    from datetime import date
    today = str(date.today())
    if os.path.exists(EMAIL_DAILY_FILE):
        try:
            with open(EMAIL_DAILY_FILE, 'r') as f:
                data = json.load(f)
            if data.get('date') == today:
                return data.get('count', 0)
        except Exception:
            pass
    return 0


def save_email_daily_count(count: int):
    """오늘 이메일 발송 건수 저장"""
    from datetime import date
    with open(EMAIL_DAILY_FILE, 'w') as f:
        json.dump({'date': str(date.today()), 'count': count}, f)


def send_emails(limit=None, test_mode=False):
    if test_mode:
        print(f"\n테스트 이메일 → {GMAIL_EMAIL}")
        try:
            server = smtplib.SMTP('smtp.gmail.com', 587)
            server.starttls()
            server.login(GMAIL_EMAIL, GMAIL_APP_PASSWORD)
            subject = "안녕하세요 마트마트 네이버 카페 광고 파트너쉽 제안드립니다. (테스트)"
            msg = _build_msg(GMAIL_EMAIL, subject, '테스트 업체', '')
            server.sendmail(GMAIL_EMAIL, GMAIL_EMAIL, msg.as_string())
            server.quit()
            print("  ✓ 테스트 발송 성공!")
        except smtplib.SMTPAuthenticationError:
            print("  ✗ Gmail 인증 실패! GMAIL_APP_PASSWORD 확인 필요")
        return 0

    prospects = load_prospects_from_excel()
    if not prospects:
        print("  발송할 업체 없음")
        return 0

    if limit:
        prospects = prospects[:limit]

    print(f"\n  이메일 대상: {len(prospects)}개 업체")
    for i, p in enumerate(prospects[:5], 1):
        print(f"    {i}. {p['company']} → {p['email']}")
    if len(prospects) > 5:
        print(f"    ... 외 {len(prospects) - 5}개")

    def _connect_smtp():
        """SMTP 연결 (재시도 포함)"""
        for attempt in range(3):
            try:
                s = smtplib.SMTP('smtp.gmail.com', 587, timeout=30)
                s.starttls()
                s.login(GMAIL_EMAIL, GMAIL_APP_PASSWORD)
                return s
            except smtplib.SMTPAuthenticationError:
                print("  ✗ Gmail 인증 실패!")
                return None
            except Exception as e:
                print(f"  ⚠ SMTP 연결 실패 (시도 {attempt+1}/3): {e}")
                if attempt < 2:
                    time.sleep(5)
        return None

    server = _connect_smtp()
    if not server:
        return 0

    subject = "안녕하세요 마트마트 네이버 카페 광고 파트너쉽 제안드립니다."
    success = 0

    # 오늘 이미 보낸 건수 확인 (재시작해도 유지)
    daily_count = load_email_daily_count()
    remaining_today = DAILY_SEND_LIMIT - daily_count
    print(f"  오늘 발송: {daily_count}건 / 한도 {DAILY_SEND_LIMIT}건 (남은 횟수: {remaining_today}건)")
    if remaining_today <= 0:
        print(f"  오늘 이메일 한도 도달. 내일 다시 실행하세요.")
        server.quit()
        return 0
    prospects = prospects[:remaining_today]

    for i, p in enumerate(prospects, 1):
        if success >= remaining_today:
            print(f"\n  이메일 일일 한도({DAILY_SEND_LIMIT}건) 도달!")
            break
        print(f"  [{i}/{len(prospects)}] {p['company']} ({p['email']})... ", end="", flush=True)
        try:
            msg = _build_msg(p['email'], subject, p['company'], p['category'])
            server.sendmail(GMAIL_EMAIL, p['email'], msg.as_string())
            mark_sent_in_excel(p['row_num'], p['email'])
            print("✓")
            success += 1
            daily_count += 1
            save_email_daily_count(daily_count)
        except smtplib.SMTPServerDisconnected:
            # 연결 끊김 → 재연결 후 재시도
            print(f"⚠ 연결 끊김, 재연결 중...", end="", flush=True)
            server = _connect_smtp()
            if not server:
                print("  ✗ 재연결 실패, 중단")
                break
            try:
                server.sendmail(GMAIL_EMAIL, p['email'], msg.as_string())
                mark_sent_in_excel(p['row_num'], p['email'])
                print("✓ (재연결 후 성공)")
                success += 1
                daily_count += 1
                save_email_daily_count(daily_count)
            except Exception as e2:
                print(f"✗ ({e2})")
        except Exception as e:
            print(f"✗ ({e})")

        if i < len(prospects) and success < remaining_today:
            delay = random.uniform(*DELAY_BETWEEN_EMAILS)
            print(f"         {delay:.0f}초 대기...", flush=True)
            time.sleep(delay)

    try:
        server.quit()
    except Exception:
        pass  # 이미 끊어진 연결 → 무시
    print(f"\n  이메일 발송 완료: {success}건")
    return success


# ============================================================
#  메인 파이프라인: 수집 → 요약 → 승인 → 발송
# ============================================================
def run_full_pipeline():
    """전체 파이프라인 (인자 없이 실행 시 기본 동작)"""

    print("\n" + "=" * 60)
    print("  마트마트 광고주 수집 & 발송 파이프라인 시작")
    print("  ★ 브라우저 안 띄움 — 백그라운드 실행 가능")
    print("=" * 60)

    # PHASE 1: 업체 수집
    collector = BusinessCollector()
    email_count = collector.collect_all()

    # PHASE 2: 인스타 수집
    print("\n" + "=" * 60)
    print(" 인스타그램 비즈니스 계정 수집")
    print("=" * 60)

    insta_accounts = {}
    insta = InstagramManager()
    if insta.login():
        insta_accounts = insta.collect_business_accounts()
        if insta_accounts:
            with open("인스타_수집결과.json", 'w', encoding='utf-8') as f:
                json.dump(insta_accounts, f, ensure_ascii=False, indent=2)

            for username, info in insta_accounts.items():
                ext_url = info.get('external_url', '')
                if ext_url:
                    for d, ci in collector.collected.items():
                        if ext_url and d in ext_url:
                            ci['instagram'] = f"@{username}"
            collector.save_to_excel()
    else:
        if os.path.exists("인스타_수집결과.json"):
            with open("인스타_수집결과.json", 'r', encoding='utf-8') as f:
                insta_accounts = json.load(f)
            print(f"  이전 인스타 수집 결과 사용: {len(insta_accounts)}개")

    # PHASE 3: 요약
    email_targets = []
    if os.path.exists(EXCEL_FILE):
        email_targets = load_prospects_from_excel()

    dm_targets = {k: v for k, v in insta_accounts.items() if k not in insta.dm_sent}
    remaining_dm = max(0, DAILY_DM_LIMIT - insta.dm_today_count)

    print("\n" + "=" * 60)
    print("  수집 완료! 발송 대상 요약")
    print("=" * 60)
    print(f"\n  총 수집 업체: {len(collector.collected)}개")
    print(f"  이메일 발송 대상: {len(email_targets)}개")
    if email_targets:
        for i, p in enumerate(email_targets[:5], 1):
            print(f"     {i}. {p['company'][:15]} → {p['email']}")
        if len(email_targets) > 5:
            print(f"     ... 외 {len(email_targets) - 5}개")

    print(f"\n  인스타 DM 대상: {len(dm_targets)}개 (오늘 잔여: {remaining_dm}개)")
    if dm_targets:
        for i, (username, info) in enumerate(list(dm_targets.items())[:5], 1):
            name = info.get('full_name', '') or username
            print(f"     {i}. @{username} ({name[:15]})")
        if len(dm_targets) > 5:
            print(f"     ... 외 {len(dm_targets) - 5}개")

    if not email_targets and not dm_targets:
        print("\n  발송할 대상이 없습니다.")
        return

    email_min = len(email_targets) * 35 // 60
    dm_actual = min(len(dm_targets), remaining_dm)
    dm_min = dm_actual * 135 // 60
    print(f"\n  예상 소요 시간: 이메일 ~{email_min}분 + DM ~{dm_min}분")

    # PHASE 4: 승인
    print(f"\n{'─' * 60}")
    print("  발송 옵션:")
    print("    YES  → 이메일 + DM 동시 발송")
    print("    E    → 이메일만 발송")
    print("    D    → DM만 발송")
    print("    NO   → 발송 안 함 (수집 결과만 저장)")
    print(f"{'─' * 60}")
    confirm = input("\n  선택: ").strip().upper()

    if confirm == 'NO' or not confirm:
        print("  발송 건너뜀. 수집 결과는 저장되어 있습니다.")
        print(f"  나중에 발송하려면: python3 마트마트_자동화.py send")
        return

    do_email = confirm in ('YES', 'E')
    do_dm = confirm in ('YES', 'D')

    # PHASE 5: 발송
    email_success = 0
    dm_success = 0

    if do_email and email_targets:
        print("\n" + "─" * 40)
        print(" 이메일 발송 시작...")
        print("─" * 40)
        email_success = send_emails()

    if do_dm and dm_targets and remaining_dm > 0:
        print("\n" + "─" * 40)
        print(" 인스타 DM 발송 시작...")
        print("─" * 40)
        if insta.client:
            dm_success = insta.send_dms_batch(dm_targets)
        elif insta.login():
            dm_success = insta.send_dms_batch(dm_targets)

    print(f"\n{'=' * 60}")
    print(f"  완료!")
    if do_email:
        print(f"  이메일: {email_success}건 발송")
    if do_dm:
        print(f"  DM: {dm_success}건 발송")
    print(f"{'=' * 60}")


def show_status():
    print("\n  [업체 수집 현황]")
    if os.path.exists(PROGRESS_FILE):
        with open(PROGRESS_FILE, 'r', encoding='utf-8') as f:
            data = json.load(f)
        collected = data.get('collected', {})
        we = sum(1 for v in collected.values() if v.get('email'))
        print(f"  총 수집: {len(collected)}개 / 이메일: {we}개")
        print(f"  네이버/구글 키워드: {len(data.get('done_keywords', []))}/{len(KEYWORDS)}개")
        print(f"  쇼핑 키워드: {len(data.get('done_shopping', []))}/{len(SHOPPING_KEYWORDS)}개")
        print(f"  지도 키워드: {len(data.get('done_map', []))}/{len(MAP_KEYWORDS)}개")
        print(f"  마지막: {data.get('last_update', '?')}")

        # 출처별 통계
        sources = {}
        for v in collected.values():
            s = v.get('source', '?')
            sources[s] = sources.get(s, 0) + 1
        print(f"  출처별: {', '.join(f'{k}={v}' for k, v in sorted(sources.items()))}")
    else:
        print("  수집 기록 없음")

    print("\n  [인스타 DM 현황]")
    if os.path.exists(INSTA_DM_PROGRESS_FILE):
        with open(INSTA_DM_PROGRESS_FILE, 'r', encoding='utf-8') as f:
            dm_data = json.load(f)
        sent = dm_data.get('sent', {})
        today = datetime.now().strftime('%Y-%m-%d')
        today_count = sum(1 for v in sent.values() if v.startswith(today))
        print(f"  총 DM: {len(sent)}개 / 오늘: {today_count}/{DAILY_DM_LIMIT}개")
    else:
        print("  DM 기록 없음")

    if os.path.exists("인스타_수집결과.json"):
        with open("인스타_수집결과.json", 'r', encoding='utf-8') as f:
            ig = json.load(f)
        print(f"  인스타 수집: {len(ig)}개 계정")


# ============================================================
#  메인 실행
# ============================================================
if __name__ == '__main__':
    print("""
╔══════════════════════════════════════════════════════════╗
║       마트마트 광고주 자동 수집 & 이메일/DM 발송         ║
║       ★ 브라우저 안 띄움! 백그라운드 실행 가능 ★        ║
╠══════════════════════════════════════════════════════════╣
║                                                          ║
║  python3 마트마트_자동화.py                               ║
║    → 전체 실행 (수집 → 요약 → 승인 → 발송)               ║
║                                                          ║
║  python3 마트마트_자동화.py collect-shopping              ║
║    → 스마트스토어만 수집 (뉴스기사 0%, 권장)               ║
║  python3 마트마트_자동화.py status    현황 확인            ║
║  python3 마트마트_자동화.py send-test 테스트 발송          ║
║                                                          ║
╚══════════════════════════════════════════════════════════╝
""")

    cmd = sys.argv[1] if len(sys.argv) > 1 else 'start'

    if cmd in ('start', 'all', 'run'):
        run_full_pipeline()

    elif cmd == 'collect':
        collector = BusinessCollector()
        collector.collect_all()

    elif cmd == 'collect-shopping':
        # 스마트스토어 전용 수집 (뉴스기사 0% — 가장 깨끗한 소스)
        # done_shopping 리셋 → 110개 신규 키워드 전체 재수집
        collector = BusinessCollector()
        collector.done_shopping = set()   # 기존 완료 키워드 초기화
        collector.save_progress()

        remaining_shop = list(SHOPPING_KEYWORDS)
        print(f"\n쇼핑 전용 수집 시작: {len(remaining_shop)}개 키워드")
        print(f"현재 누적: {len(collector.collected)}개\n")
        before = len(collector.collected)
        import random, time
        for i, kw in enumerate(remaining_shop, 1):
            print(f"  ({i}/{len(remaining_shop)}) '{kw}'", end="", flush=True)
            n = collector.search_naver_shopping(kw)
            print(f" → +{n}개 (누적 {len(collector.collected)})")
            collector.done_shopping.add(kw)
            if i % 10 == 0:
                collector.save_progress()
            time.sleep(random.uniform(1, 2))
        collector.save_progress()
        added = len(collector.collected) - before
        print(f"\n쇼핑 수집 완료: +{added}개 신규 (총 {len(collector.collected)}개)")
        print("→ 이메일 보강은 'python3 마트마트_자동화.py collect' 전체 실행 시 자동 처리")

    elif cmd == 'collect-insta':
        insta = InstagramManager()
        if insta.login():
            accounts = insta.collect_business_accounts()
            if accounts:
                with open("인스타_수집결과.json", 'w', encoding='utf-8') as f:
                    json.dump(accounts, f, ensure_ascii=False, indent=2)
                print(f"\n  저장 완료: {len(accounts)}개 계정")

    elif cmd == 'send-test':
        send_emails(test_mode=True)

    elif cmd == 'send-dm':
        # 인스타_수집결과.json에서 광고적합=True 계정 로드 후 DM 발송
        insta_file = os.path.join(os.path.dirname(os.path.abspath(__file__)), '인스타_수집결과.json')
        if not os.path.exists(insta_file):
            print("  ✗ 인스타_수집결과.json 파일 없음")
            sys.exit(1)
        with open(insta_file, 'r', encoding='utf-8') as f:
            all_accounts = json.load(f)
        targets = {u: v for u, v in all_accounts.items() if v.get('광고적합') == True and v.get('user_id')}
        print(f"\n  DM 대상: {len(targets)}개 (광고적합 계정)")
        insta = InstagramManager()
        if insta.login():
            insta.send_dms_batch(targets)
        else:
            print("  ✗ 인스타그램 로그인 실패")

    elif cmd == 'send':
        limit = int(sys.argv[2]) if len(sys.argv) > 2 else None
        run_full_pipeline()

    elif cmd == 'send-all':
        import threading, subprocess as sp

        def run_email():
            print("\n[이메일] 발송 시작")
            send_emails()

        def run_dm():
            dm_script = os.path.join(os.path.dirname(os.path.abspath(__file__)), '인스타_DM_자동화.py')
            if not os.path.exists(dm_script):
                print("  [DM] ✗ 인스타_DM_자동화.py 없음")
                return
            print("\n[DM] 브라우저 실행 중...")
            sp.run([sys.executable, dm_script])

        t_email = threading.Thread(target=run_email, name="이메일")
        t_dm    = threading.Thread(target=run_dm,    name="인스타DM")
        t_email.start()
        t_dm.start()
        t_email.join()
        t_dm.join()
        print("\n\n✅ 이메일 + 인스타 DM 발송 완료")

    elif cmd == 'status':
        show_status()

    else:
        print(f"  알 수 없는 명령: {cmd}")
        print("  그냥 python3 마트마트_자동화.py 로 실행하세요.")
